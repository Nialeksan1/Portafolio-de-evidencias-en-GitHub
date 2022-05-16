# Autores: Perla Marlene Viera Gonzáles
# Modificado por: Nineck-Aleksander Hournou 1913747

from pyhunter import PyHunter
from openpyxl import Workbook
import getpass


def Busqueda(organizacion):
    # Cantidad de resultados esperados de la búsqueda
    # El límite MENSUAL de Hunter es 50, cuidado!
    resultado = hunter.domain_search(company=organizacion, limit=1, emails_type='personal')
    return resultado


def GuardarInformacion(datos_encontrados, organizacion):
    # Se crea un libro y una hoja de cálculo en Excel
    libro = Workbook()
    libro.create_sheet(organizacion)
    libro.save("Hunter" + organizacion + ".xlsx")
    libro.active = libro.sheetnames.index(organizacion)
    hoja_activa = libro.active
    # Se modifican dos filas y dos columnas con los valores del diccionario
    hoja_activa['A1'] = 'Domain ->'
    hoja_activa['B1'] = datos_encontrados['domain']
    hoja_activa['A2'] = 'Organization ->'
    hoja_activa['B2'] = datos_encontrados['organization']
    hoja_activa['A3'] = 'Email ->'
    # Se crea un ciclo para entrar a la lista
    row = 3
    for x in datos_encontrados['emails']:
        # Una vez dentro de la lista, se crea un ciclo para entrar al diccionario
        for j, k in x.items():
            if j == 'first_name':
                break
            # Se hace el mismo procedimiento de arriba, solo que para los valores de 'sources'
            if j == 'sources':
                hoja_activa.cell(row, 2, j + ' ->')
                for y in k:
                    for l, m in y.items():
                        hoja_activa.cell(row, 3, l + ' ->')
                        hoja_activa.cell(row, 4, m)
                        row += 1
                    hoja_activa.cell(row, 3, None)
                    hoja_activa.cell(row, 4, None)
                    row += 1
            else:
                hoja_activa.cell(row, 2, j + ' ->')
                hoja_activa.cell(row, 3, k)
            row += 1
    libro.save("Hunter" + organizacion + ".xlsx")

# Se pide la llave y el dominio al usuario
print("Script para buscar información")
apikey = getpass.getpass("Ingresa tu API key: ")
hunter = PyHunter(apikey)
orga = input("Dominio a investigar: ")
datos_encontrados = Busqueda(orga)
# Si no se encontró información, sale del programa, sino, continúa con los procesos
if datos_encontrados is None:
    exit()
else:
    print(datos_encontrados)
    print(type(datos_encontrados))
    GuardarInformacion(datos_encontrados, orga)
